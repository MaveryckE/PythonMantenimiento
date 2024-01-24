from conf.conexion2 import conn
from fastapi import FastAPI, Response, APIRouter
from pydantic import BaseModel
from typing import List
from starlette.status import HTTP_400_BAD_REQUEST, HTTP_200_OK
from openpyxl import Workbook

class listar(BaseModel):
     categoria: str
     descripcion: str
     
class productosbase(BaseModel):
    codigo_producto:int
    nombre_producto:str
    descripcion:str 
    precio:int
    tipoproducto:str
     
class proveedoresbase(BaseModel):
    codigoproveedor:int
    dpiproveedor: int
    nombreproveedor:str
    apellidoproveedor:str
    correoproveedor:str
    direccion:str
    telefono:int  

mods = APIRouter()

@mods.get('/api/proceso', response_model= List[listar])
def prueba():
    cursor1 = conn.cursor()
    result = cursor1.execute("SELECT * FROM CATEGORIAS").fetchall()
    print(result)
    item = []
    for fila in result:
        patito={}
        patito["categoria"]=fila[1]
        patito["descripcion"]=fila[2]
        item.append(patito)

    return item

#GENERACION DE LIBROS DE COMPRAS 
class Compra(BaseModel):
    fecha: str
    proveedor: str
    descripcion: str
    monto: float

 

@mods.get('/libro-de-compras', response_model=List[Compra])
def generar_libro_de_compras():
    cursor = conn.cursor()
    result = cursor.execute("SELECT fecha, proveedor, descripcion, monto FROM compras").fetchall()

    libro_de_compras = []
    for fila in result:
        compra = Compra(fecha=fila[0], proveedor=fila[1], descripcion=fila[2], monto=fila[3])
        libro_de_compras.append(compra)

    return libro_de_compras

    #inicio mantenenimiento
    
@mods.get("/api/verProveedores", response_model=List[proveedoresbase])
def verProveedores():
    cursor = conn.cursor()
    resultado = cursor.execute("SELECT * FROM PROVEEDOR").fetchall()
    proveedores = []
    for re in resultado:
        llenar ={}
        llenar["codigoproveedor"]= re[0]
        llenar["dpiproveedor"]=re[1]
        llenar["nombreproveedor"]= re[2]
        llenar["apellidoproveedor"]= re[3]
        llenar["correoproveedor"]= re[4]
        llenar["direccion"]= re[5]
        llenar["telefono"]= re[6]
        proveedores.append(llenar)
    return proveedores  

@mods.post("/api/AgregarProveedor")
def AgregarProveedor(lista:proveedoresbase):
    cursor = conn.cursor()
    resultado = cursor.execute("INSERT INTO PROVEEDOR(DPI_PROVEEDOR, NOMBRE_PROVEEDOR, APELLIDO_PROVEEDOR, CORREO_PROVEEDOR, DIRECCION_PROVEEDOR, TELEFONO_PROVEEDOR) VALUES (:1, :2, :3, :4, :5, :6)", (
                                lista.dpiproveedor,
                                lista.nombreproveedor,
                                lista.apellidoproveedor,
                                lista.correoproveedor,
                                lista.direccion,
                                lista.telefono
                            ))
    conn.commit()
    return Response(status_code=HTTP_200_OK)

@mods.put("/api/ModificarProveedor")
def ModificarProveedor(lista: proveedoresbase):
    cursor = conn.cursor()
    resultado = cursor.execute("UPDATE PROVEEDOR SET DPI_PROVEEDOR =:1, NOMBRE_PROVEEDOR=:2, APELLIDO_PROVEEDOR=:3, CORREO_PROVEEDOR=:4, DIRECCION_PROVEEDOR=:5,\
                                TELEFONO_PROVEEDOR=:6 WHERE CODIGO_PROVEEDOR =:7", (
                                lista.dpiproveedor,
                                lista.nombreproveedor,
                                lista.apellidoproveedor,
                                lista.correoproveedor,
                                lista.direccion,
                                lista.telefono,
                                lista.codigoproveedor,
                                ))
    conn.commit()
    return Response(status_code=HTTP_200_OK)

@mods.delete("/api/EliminarProveedor")
def EliminarProveedor(lista: proveedoresbase):
    cursor = conn.cursor()
    resultado = cursor.execute("DELETE PROVEEDOR WHERE CODIGO_PROVEEDOR ="+ str(lista.codigoproveedor))
    conn.commit()
    return Response(status_code=HTTP_200_OK)
#SELECT * FROM  PROVEEDOR WHERE CODIGO_PROVEEDOR = 
@mods.get("/api/buscarProv/<id>", response_model=proveedoresbase)
def buscarProv(id : str):
    cursor = conn.cursor()
    resultado = cursor.execute("SELECT * FROM  PROVEEDOR WHERE CODIGO_PROVEEDOR = "+id).fetchone()
  
    llenar ={
    "codigoproveedor": int(resultado[0]),
    "dpiproveedor":int(resultado[1]),
    "nombreproveedor": resultado[2],
    "apellidoproveedor": resultado[3],
    "correoproveedor": resultado[4],
    "direccion": resultado[5],
    "telefono": int(resultado[6])      
    }
    return llenar  

#productos

@mods.get("/api/verProductos", response_model=List[productosbase])
def verProductos():
    cursor = conn.cursor()
    resultado = cursor.execute("SELECT * FROM PRODUCTO").fetchall()
    producto = []
    for re in resultado:
        llenar ={}
        llenar["codigo_producto"]=re[0]
        llenar["nombre_producto"]=re[1]
        llenar["descripcion"]=re[2] 
        llenar["precio"]=re[3]
        llenar["tipoproducto"]=re[4]
        producto.append(llenar)
    return producto

@mods.post("/api/CreaProducto")
def crearProducto(lista: productosbase):
    cursor = conn.cursor()
    cursor.execute("INSERT INTO PRODUCTO(NOMBRE_PRODUCTO, DESCRIPCION_PRODUCTO, PRECIO_PRODUCTO, TIPO_PRODUCTO)\
                    VALUES(:1,:2,:3,:4)", (
                        lista.nombre_producto,
                        lista.descripcion,
                        lista.precio,
                        lista.tipoproducto
                    ))
    conn.commit()
    return Response(status_code=HTTP_200_OK)

@mods.put("/api/modificarProducto")
def modificarProducto(lista: productosbase):
    cursor = conn.cursor()
    cursor.execute("UPDATE PRODUCTO SET NOMBRE_PRODUCTO =:1, DESCRIPCION_PRODUCTO =:2, PRECIO_PRODUCTO=:3, TIPO_PRODUCTO =:4\
                    WHERE CODIGO_PRODUCTO =:5", (
                        lista.nombre_producto,
                        lista.descripcion,
                        lista.precio,
                        lista.tipoproducto,
                        lista.codigo_producto
                    ))
    conn.commit()
    return Response(status_code=HTTP_200_OK)

@mods.delete("/api/EliminarProducto")
def EliminarProducto(lista: productosbase):
    cursor = conn.cursor()
    cursor.execute("DELETE PRODUCTO WHERE CODIGO_PRODUCTO ="+ str(lista.codigo_producto))
    conn.commit()
    return Response(status_code=HTTP_200_OK)

@mods.get("/api/buscarProducto/<id>", response_model=productosbase)
def buscarProducto(id:str):
    cursor = conn.cursor()
    resultado = cursor.execute("SELECT * FROM PRODUCTO WHERE CODIGO_PRODUCTO = "+ id).fetchone()
   
    llenar = {
        "codigo_producto": resultado[0],
        "nombre_producto": resultado[1],
        "descripcion": resultado[2], 
        "precio": resultado[3],
        "tipoproducto": resultado[4],   
     }
        
    return llenar


@mods.get("/api/exportarDatosAExcelll")
def exportarDatoAExcelll():
    cursor = conn.cursor()

    #consulta para exportar los datos 
    sql_query = "SELECT CODIGO_PRODUCTO, NOMBRE_PRODUCTO, DESCRIPCION_PRODUCTO, PRECIO_PRODUCTO, TIPO_PRODUCTO FROM PRODUCTO"
    cursor.execute(sql_query)
    datos = cursor.fetchall()

    # Crea un nuevo libro de Excel
    libro_excel = Workbook()
    hoja = libro_excel.active

    # Agrega encabezados a la hoja de Excel
    hoja['A1'] = "CODIGO DE PRODUCTOS"
    hoja['B1'] = "NOMBRE DE PRODUCTO"
    hoja['C1'] = "DESCRIPCION DE PRODUCTO"
    hoja['D1'] = "PRECIO DE PRODUCTO"
    hoja['E1'] = "TIPO DE PRODUCTO"

    # Itera a través de los datos y agrégalos a la hoja de Excel
    fila = 2
    for registro in datos:
        hoja.cell(row=fila, column=1, value=registro[0])
        hoja.cell(row=fila, column=2, value=registro[1])
        hoja.cell(row=fila, column=3, value=registro[2])
        hoja.cell(row=fila, column=4, value=registro[3])
        hoja.cell(row=fila, column=5, value=registro[4])
        fila += 1

    # Guarda el libro de Excel en un archivo
    archivo_excel = "datos_exportados.xlsx"
    libro_excel.save(archivo_excel)

    # Cierra el cursor y la conexión a la base de datos Oracle
    cursor.close()
    conn.close()

    # Devuelve el archivo Excel como respuesta
    with open(archivo_excel, "rb") as excel_file:
        response = Response(content=excel_file.read(), status_code=HTTP_200_OK)
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = f"attachment; filename={archivo_excel}"

    return response

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(mods, host="127.0.0.1", port=8000)
    
    
@mods.get("/api/excelProveedores")
def excelProveedores():
    cursor = conn.cursor()

    sql_query = "SELECT CODIGO_PROVEEDOR, DPI_PROVEEDOR, NOMBRE_PROVEEDOR, APELLIDO_PROVEEDOR, CORREO_PROVEEDOR, DIRECCION_PROVEEDOR, TELEFONO_PROVEEDOR FROM PROVEEDOR"
    cursor.execute(sql_query)
    datos = cursor.fetchall()

    libro_excel = Workbook()
    hoja = libro_excel.active

    hoja['A1'] = "CODIGO PROVEEDOR"
    hoja['B1'] = "DPI PROVEEDOR PROVEEDOR"
    hoja['C1'] = "NOMBRE PROVEEDOR"
    hoja['D1'] = "APELLIDO PROVEEDOR"
    hoja['E1'] = "CORREO PROVEEDOR"
    hoja['F1'] = "DIRECCION PROVEEDOR"
    hoja['G1'] = "TELEFONO PROVEEDOR"

    fila = 2
    for registro in datos:
        hoja.cell(row=fila, column=1, value=registro[0])
        hoja.cell(row=fila, column=2, value=registro[1])
        hoja.cell(row=fila, column=3, value=registro[2])
        hoja.cell(row=fila, column=4, value=registro[3])
        hoja.cell(row=fila, column=5, value=registro[4])
        hoja.cell(row=fila, column=6, value=registro[5])
        hoja.cell(row=fila, column=7, value=registro[6])
        fila += 1

    archivo_excel = "C:\\Users\juego\\Downloads\\MISEXCEL\\excelProveedores.xlsx"
    libro_excel.save(archivo_excel)

    return Response(status_code=HTTP_200_OK)
    
#PROCESO DOS
@mods.get("/api/procesosuma", response_model= List[productosbase])
def procesosuma():
    cursor1 = conn.cursor()
    result = cursor1.execute("SELECT SUM(PRECIO_PRODUCTO) FROM PRODUCTO;").fetchone()
    print(result)
    item = []
    for fila in result:
        patito={}
        patito["codigo_producto"]=fila[0]
        patito["nombre_producto"]=fila[1]
        patito["descripcion"]=fila[2]
        patito["precio"]=fila[3]
        patito["tipoproducto"]=fila[4]
        item.append(patito)
        
    return item